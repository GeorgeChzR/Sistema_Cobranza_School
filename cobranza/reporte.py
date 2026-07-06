"""Exportación del análisis de cartera a un reporte Excel con formato."""

from __future__ import annotations

import io
from pathlib import Path

import pandas as pd

from .analisis import resumen_indicadores

COLORES_FONDO = {
    "Verde": "C6EFCE",
    "Amarillo": "FFEB9C",
    "Naranja": "FCD5B4",
    "Rojo": "FFC7CE",
}
COLORES_TEXTO = {
    "Verde": "006100",
    "Amarillo": "9C6500",
    "Naranja": "974706",
    "Rojo": "9C0006",
}


def exportar_reporte(
    estado: pd.DataFrame,
    ruta_salida: str | Path,
    fecha_corte=None,
) -> Path:
    """Escribe un Excel en disco (uso CLI)."""

    ruta_salida = Path(ruta_salida)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    ruta_salida.write_bytes(exportar_reporte_bytes(estado, fecha_corte))
    return ruta_salida


def exportar_reporte_bytes(estado: pd.DataFrame, fecha_corte=None) -> bytes:
    """Genera el Excel en memoria, sin crear archivos temporales en disco."""

    buffer = io.BytesIO()
    indicadores = resumen_indicadores(estado, fecha_corte)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _escribir_hojas(writer, estado, indicadores)

    buffer.seek(0)
    from openpyxl import load_workbook

    wb = load_workbook(buffer)
    _aplicar_formato_workbook(wb)
    salida = io.BytesIO()
    wb.save(salida)
    return salida.getvalue()


def _escribir_hojas(writer, estado: pd.DataFrame, indicadores: dict) -> None:
    _hoja_resumen(writer, indicadores)
    _hoja_datos(writer, estado, "Todos")

    for color in ["Rojo", "Naranja", "Amarillo", "Verde"]:
        sub = estado[estado["Semáforo"] == color]
        if not sub.empty:
            _hoja_datos(writer, sub, color)

    avisar = estado[estado["Saldo"] > 0.005].copy()
    _hoja_datos(writer, avisar, "A_Contactar")


def _hoja_resumen(writer, ind: dict) -> None:
    if not ind or ind.get("total_alumnos", 0) == 0:
        pd.DataFrame([{"Mensaje": "Sin datos"}]).to_excel(
            writer, sheet_name="Resumen", index=False
        )
        return

    filas = [
        ["Fecha de corte", ind.get("fecha_corte")],
        ["Total de alumnos", ind.get("total_alumnos")],
        ["Alumnos con adeudo vencido", f"{ind.get('alumnos_con_adeudo_vencido', 0)} ({ind.get('porcentaje_adeudo_vencido', 0)}%)"],
        ["Saldo total pendiente", ind.get("saldo_total")],
        ["", ""],
        ["Semáforo", "Alumnos", "% Cartera", "Saldo"],
    ]
    for c in ["Verde", "Amarillo", "Naranja", "Rojo"]:
        filas.append(
            [
                c,
                ind["conteo"].get(c, 0),
                ind["porcentaje"].get(c, 0),
                ind["saldo_por_color"].get(c, 0),
            ]
        )
    pd.DataFrame(filas).to_excel(
        writer, sheet_name="Resumen", index=False, header=False
    )


def _hoja_datos(writer, df: pd.DataFrame, nombre: str) -> None:
    if df is None or df.empty:
        pd.DataFrame([{"Mensaje": "Sin registros"}]).to_excel(
            writer, sheet_name=nombre[:31], index=False
        )
        return
    df.to_excel(writer, sheet_name=nombre[:31], index=False)


def _aplicar_formato_workbook(wb) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    encabezado_fill = PatternFill("solid", fgColor="305496")
    encabezado_font = Font(bold=True, color="FFFFFF")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = encabezado_fill
            cell.font = encabezado_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        col_sem = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip() == "Semáforo":
                col_sem = idx
                break

        if col_sem:
            for fila in range(2, ws.max_row + 1):
                valor = ws.cell(row=fila, column=col_sem).value
                if valor in COLORES_FONDO:
                    ws.cell(row=fila, column=col_sem).fill = PatternFill(
                        "solid", fgColor=COLORES_FONDO[valor]
                    )
                    ws.cell(row=fila, column=col_sem).font = Font(
                        bold=True, color=COLORES_TEXTO[valor]
                    )

        for col in range(1, ws.max_column + 1):
            letra = get_column_letter(col)
            largo = max(
                (
                    len(str(ws.cell(row=r, column=col).value or ""))
                    for r in range(1, min(ws.max_row, 200) + 1)
                ),
                default=10,
            )
            ws.column_dimensions[letra].width = min(max(largo + 2, 12), 45)

        ws.freeze_panes = "A2"
